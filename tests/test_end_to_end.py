"""Integration tests verifying the QTO refactor — no Vision/AI calls needed."""
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))


# ── Sheet number normalization ───────────────────────────────────────────────

def test_sheet_normalization():
    from parser.title_block_reader import normalize_sheet_number
    assert normalize_sheet_number("A107.00") == "A-107"
    assert normalize_sheet_number("A106.00") == "A-106"
    assert normalize_sheet_number("A901.00") == "A-901"
    assert normalize_sheet_number("A-106") == "A-106"
    assert normalize_sheet_number("R-001") == "R-001"
    assert normalize_sheet_number("T-002") == "T-002"


def test_sheet_normalization_no_suffix():
    from parser.title_block_reader import normalize_sheet_number
    assert ".00" not in normalize_sheet_number("A107.00")
    assert ".00" not in normalize_sheet_number("A106.00")


# ── Description normalizer examples ─────────────────────────────────────────

def test_description_examples():
    from ai.description_composer import _SYSTEM
    # Real reference examples must be present
    assert "MAPLE FLOORING (25/32\" T)" in _SYSTEM
    assert "SBS MODIFIED" in _SYSTEM
    assert "(1'-3\" T) CAST STONE COPING" in _SYSTEM
    assert "(ALLOWANCE)" in _SYSTEM
    assert "(PROVISION)" in _SYSTEM
    assert "WHICH INCLUDES" in _SYSTEM
    assert "SQFT)" in _SYSTEM  # math trail
    assert "@ AUDITORIUM" in _SYSTEM
    assert "@ ROOF" in _SYSTEM
    assert "LEGEND/A102" in _SYSTEM


def test_description_examples_count():
    from ai.description_composer import _SYSTEM
    # Count Input:/Output: pairs — must have at least 15 examples
    input_count = _SYSTEM.count("\nInput:")
    assert input_count >= 15, f"Only {input_count} examples found; need ≥15"


def test_description_no_invented_examples():
    from ai.description_composer import _SYSTEM
    # Original 8 invented examples must not appear
    invented = [
        "Remove and replace brick veneer at parapet, approx 15LF",
        "Install new EPDM roofing membrane over existing substrate",
        "Repoint mortar joints at masonry wall, 8th floor corridor",
        "Paint existing exposed ductwork and conduit",
        "Remove existing wood window and install new aluminum window",
        "Coping replacement at roof parapet, cast stone, approx 180 LF",
        "Clean existing masonry facade, apply waterproof sealer",
        "Patch plaster ceiling where damaged, match existing texture",
    ]
    for snippet in invented:
        assert snippet not in _SYSTEM, f"Invented example still present: {snippet!r}"


# ── Assembler routing ────────────────────────────────────────────────────────

def test_assembler_legend_routing():
    """Legend extractor must be called for non-T-002 pages."""
    from core.assembler import Assembler
    config = {"extraction_mode": "standard"}
    mock_ai = MagicMock()
    mock_tracker = MagicMock()

    with patch("core.assembler.read_title_block") as mock_tb, \
         patch("core.assembler.detect_tables", return_value=[]), \
         patch("core.assembler.detect_scale", return_value=None), \
         patch("core.assembler.extract_legend_items", return_value=[]) as mock_legend, \
         patch("core.assembler.extract_allowances") as mock_allowance:
        from parser.title_block_reader import TitleBlockInfo
        mock_tb.return_value = TitleBlockInfo(sheet_number="A-106")
        from parser.pdf_splitter import PageInfo
        page_info = PageInfo(page_num=1, page_type="PLAN_CONSTRUCTION", text="", skip=False)
        page = MagicMock()
        assembler = Assembler(config, mock_ai, mock_tracker)
        assembler.process_page(page, page_info, "test.pdf")
        mock_legend.assert_called_once()
        mock_allowance.assert_not_called()


def test_assembler_allowance_routing():
    """Allowance extractor must be called for T-002 pages."""
    from core.assembler import Assembler
    config = {"extraction_mode": "standard"}
    mock_ai = MagicMock()
    mock_tracker = MagicMock()

    with patch("core.assembler.read_title_block") as mock_tb, \
         patch("core.assembler.extract_allowances", return_value=[]) as mock_allowance, \
         patch("core.assembler.extract_legend_items") as mock_legend:
        from parser.title_block_reader import TitleBlockInfo
        mock_tb.return_value = TitleBlockInfo(sheet_number="T-002")
        from parser.pdf_splitter import PageInfo
        page_info = PageInfo(page_num=1, page_type="SCHEDULE", text="", skip=False)
        page = MagicMock()
        assembler = Assembler(config, mock_ai, mock_tracker)
        assembler.process_page(page, page_info, "test.pdf")
        mock_allowance.assert_called_once()
        mock_legend.assert_not_called()


# ── XLSX exporter ────────────────────────────────────────────────────────────

def _make_mock_ws_with_chain_formulas():
    """Create an in-memory workbook with column A chain formulas in rows 44-99."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A43"] = 43
    for r in range(44, 100):
        ws.cell(row=r, column=1).value = f"=A{r - 1}+1"
        ws.cell(row=r, column=2).value = f"B{r} data"
    return ws


def test_exporter_chain_formula():
    """Column A chain formulas must survive _clear_preallocated."""
    from core.xlsx_exporter import _clear_preallocated
    ws = _make_mock_ws_with_chain_formulas()
    _clear_preallocated(ws)
    # All column A formulas should still be present
    for r in range(44, 100):
        val = ws.cell(row=r, column=1).value
        assert val == f"=A{r - 1}+1", f"Row {r} col A was cleared: {val!r}"
    # Column B should be cleared
    for r in range(44, 100):
        assert ws.cell(row=r, column=2).value is None, f"Row {r} col B not cleared"


def test_exporter_wrap_text():
    """D column must have wrap_text=True after _write_data_row."""
    from core.xlsx_exporter import _write_data_row
    from core.qto_row import QTORow
    wb = openpyxl.Workbook()
    ws = wb.active
    row = QTORow(
        description="LINE ONE\n-SUB ITEM A\n-SUB ITEM B",
        drawings="A-106",
        details="LEGEND/A106",
        qty=10.0,
        units="LF",
    )
    _write_data_row(ws, 44, row)
    assert ws["D44"].alignment.wrap_text is True


def test_exporter_row_height():
    """Row height must be proportional to newline count."""
    from core.xlsx_exporter import _write_data_row
    from core.qto_row import QTORow
    wb = openpyxl.Workbook()
    ws = wb.active
    row = QTORow(
        description="LINE ONE\n-A\n-B\n-C",
        drawings="A-106",
        details="LEGEND/A106",
        qty=5.0,
        units="SQ FT",
    )
    _write_data_row(ws, 44, row)
    expected_height = 15 * max(2, 4)  # 4 lines → 60
    assert ws.row_dimensions[44].height == expected_height
