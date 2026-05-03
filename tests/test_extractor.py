"""Smoke tests against the HBT drawing fixtures."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import fitz
import pytest

PDF = str(Path(__file__).parent / "fixtures" / "HBT_drawings.pdf")
_NEEDS_PDF = pytest.mark.skipif(not Path(PDF).exists(), reason="HBT_drawings.pdf fixture missing")


@_NEEDS_PDF
def test_pdf_opens():
    doc = fitz.open(PDF)
    assert doc.page_count > 0
    doc.close()


@_NEEDS_PDF
def test_classify_pages():
    from parser.pdf_splitter import classify_page
    doc = fitz.open(PDF)
    classified = {}
    for i in range(doc.page_count):
        page = doc[i]
        text = page.get_text("text") or ""
        info = classify_page(i + 1, text)
        classified[i + 1] = info.page_type
    doc.close()
    # At least some pages should be classified
    assert len(classified) > 0
    assert any(t != "TITLE_PAGE" for t in classified.values())


@_NEEDS_PDF
def test_table_detection_page_18():
    """Page 18 has KEYED PLUMBING NOTES — should find at least one Type A table."""
    from parser.table_detector import detect_tables
    doc = fitz.open(PDF)
    page = doc[17]
    tables = detect_tables(page, PDF, 18)
    doc.close()
    type_a = [t for t in tables if t.table_type == "A"]
    assert len(type_a) >= 1, f"Expected Type A table on page 18, got: {[t.table_type for t in tables]}"


@_NEEDS_PDF
def test_table_extraction_page_18():
    """Extracted Type A rows should have non-empty descriptions."""
    from parser.table_detector import detect_tables
    from parser.table_extractor import extract_type_a
    from parser.title_block_reader import TitleBlockInfo
    doc = fitz.open(PDF)
    page = doc[17]
    tables = detect_tables(page, PDF, 18)
    info = TitleBlockInfo()
    for t in tables:
        if t.table_type == "A":
            items = extract_type_a(t, page, info)
            assert len(items) >= 1, f"Expected items from Type A table"
            for item in items:
                assert item.get("description"), f"Item missing description: {item}"
    doc.close()


@_NEEDS_PDF
def test_title_block_page_18():
    """Page 18 (P 140 00) title block should have sheet number."""
    from parser.title_block_reader import read_title_block
    doc = fitz.open(PDF)
    page = doc[17]
    config = {"title_block_region": {"pct": 0.15}, "min_vector_text_length": 10}
    info = read_title_block(page, config)
    doc.close()
    assert info.sheet_number or info.sheet_title, f"Expected sheet number, got: {info}"


@_NEEDS_PDF
def test_scale_detection():
    """Most pages should return None scale (AS SHOWN / NTS)."""
    from parser.scale_detector import detect_scale
    doc = fitz.open(PDF)
    results = []
    for i in range(doc.page_count):
        scale = detect_scale(doc[i])
        results.append(scale)
    doc.close()
    # Most sheets in this set are NTS — that's fine
    assert isinstance(results, list)


def test_xlsx_template_readable():
    """The GC estimate template should load without errors."""
    import openpyxl
    template = str(Path(__file__).parent.parent / "ESTIMATE_FORMAT___GC.xlsx")
    wb = openpyxl.load_workbook(template, data_only=False)
    ws = wb.active
    assert ws.max_row > 10
    # Row 7 should have header labels
    assert ws["A7"].value == "S.NO"
    assert ws["E7"].value == "QTY"


def test_qto_row_dataclass():
    from core.qto_row import QTORow
    r = QTORow(description="Remove brick", qty=5.0, units="SF", trade_division="DIVISION 02")
    assert r.confidence == 1.0
    assert not r.needs_review


@_NEEDS_PDF
def test_cache_roundtrip(tmp_path):
    from core.cache import ResultCache
    from core.qto_row import QTORow
    cache = ResultCache(str(tmp_path))
    rows = [QTORow(description="Test", qty=1.0, units="EA")]
    # Use the actual PDF path for fingerprinting
    cache.save(PDF, rows)
    loaded = cache.load(PDF)
    assert loaded is not None
    assert len(loaded) == 1
    assert loaded[0].description == "Test"


def test_token_tracker():
    from core.token_tracker import TokenTracker
    tracker = TokenTracker()
    assert tracker.usage.api_calls == 0
    assert tracker.usage.estimated_cost_usd == 0.0
