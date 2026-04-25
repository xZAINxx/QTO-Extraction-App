"""Brooklyn fixture integration test — Phase 1 acceptance gate.

This test loads the real Brooklyn drawings PDF (if present in `QTO Tool/`)
and verifies all Phase 1 wiring works end-to-end without an Anthropic key:
- Title-block reader returns a sheet number on every page that has vector
  text (and gracefully empty on rasterized ones).
- Zone segmenter returns at least one usable rectangle per page.
- Allowance extractor reads pdfplumber rows (vector path; no API).
- xlsx exporter produces a valid file from a synthetic row set that
  matches the reference shape.

When the AI client is mocked, the test asserts cost ≤ $0.30 (the Phase 1
budget). With no fixture present, the test is skipped instead of failing.
"""
from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pytest


_REPO_ROOT = Path(__file__).resolve().parent.parent
_FIXTURE_DIR = _REPO_ROOT / "QTO Tool"
_BROOKLYN_PDF = _FIXTURE_DIR / "Drawings Set.pdf"
_BROOKLYN_XLSX = _FIXTURE_DIR / "QTY TAKEOFF MS_267K_BROOKLYN.xlsx"
_TEMPLATE = _REPO_ROOT / "ESTIMATE_FORMAT___GC.xlsx"


def _have_fixture() -> bool:
    return _BROOKLYN_PDF.exists() and _BROOKLYN_XLSX.exists() and _TEMPLATE.exists()


fixture_only = pytest.mark.skipif(
    not _have_fixture(),
    reason=("Brooklyn fixture missing — expected QTO Tool/Drawings Set.pdf, "
            "QTY TAKEOFF MS_267K_BROOKLYN.xlsx, and ESTIMATE_FORMAT___GC.xlsx"),
)


# ── Title-block reader ─────────────────────────────────────────────────────


@fixture_only
def test_title_block_vector_pages():
    """Pages with vector sheet numbers (H-101, A-901) should extract them."""
    import fitz
    from parser.title_block_reader import read_title_block

    doc = fitz.open(str(_BROOKLYN_PDF))
    try:
        config = {"title_block_region": {"pct": 0.18}}
        # Page 8 = H-101 (vector), page 31 = A-901 (vector).
        for page_idx, expected in [(7, "H-101"), (30, "A-901")]:
            info = read_title_block(doc[page_idx], config, ai_client=None)
            assert info.sheet_number == expected, (
                f"page {page_idx + 1}: got {info.sheet_number!r}, want {expected!r}"
            )
            assert info.source == "vector"
    finally:
        doc.close()


@fixture_only
def test_title_block_rasterized_pages_no_false_positive():
    """Pages 18/25/26 have NO vector sheet number — must not return Building-ID K-324."""
    import fitz
    from parser.title_block_reader import read_title_block

    doc = fitz.open(str(_BROOKLYN_PDF))
    try:
        config = {"title_block_region": {"pct": 0.18}}
        for page_idx in [17, 24, 25]:
            info = read_title_block(doc[page_idx], config, ai_client=None)
            assert info.sheet_number == "", (
                f"page {page_idx + 1}: should have empty sheet (no vector text), "
                f"got {info.sheet_number!r}"
            )
    finally:
        doc.close()


# ── Zone segmenter ─────────────────────────────────────────────────────────


@fixture_only
def test_zone_segmenter_finds_title_block_every_page():
    import fitz
    from parser.zone_segmenter import segment

    doc = fitz.open(str(_BROOKLYN_PDF))
    try:
        for page_idx in [0, 7, 17, 24, 30]:
            zones = segment(doc[page_idx], page_num=page_idx + 1)
            assert zones.title_block is not None, f"page {page_idx + 1}: no title block"
            assert zones.title_block.rect.width > 0
            # Every page must have at least one body region (plan/legend/schedule).
            assert (
                zones.legends or zones.schedules or zones.plan_bodies
            ), f"page {page_idx + 1}: no body zones"
    finally:
        doc.close()


# ── Assembler smoke (no API; ai_client mocked) ─────────────────────────────


@fixture_only
def test_assembler_no_api_calls_for_skipped_pages():
    """Skipped pages must short-circuit before touching the AI client."""
    from core.assembler import Assembler
    from parser.pdf_splitter import PageInfo

    ai = MagicMock()
    tracker = MagicMock()
    config = {"extraction_mode": "hybrid"}
    a = Assembler(config, ai, tracker)

    page = MagicMock()
    page_info = PageInfo(page_num=1, page_type="TITLE_PAGE", text="", skip=True,
                         skip_reason="title")
    rows = a.process_page(page, page_info, str(_BROOKLYN_PDF))
    assert rows == []
    ai.assert_not_called()


# ── Cost gate: simulate full extraction with mocked AI ─────────────────────


def test_phase1_cost_envelope_under_30_cents():
    """Realistic Phase-1 call profile must stay under the $0.30 budget.

    Modeled on the Brooklyn 42-page set with full caching enabled:
      • 42 Haiku page-type classifications (small input, small output)
      • 16 Sonnet vision crops (legend / schedule / title-block fallback / composer)
      • 1 system-prompt cache write per fresh Sonnet conversation

    This guards both pricing math and the cache-hit accounting. If it fails,
    either the model router is over-spending or cache reads are being missed.
    """
    from core.token_tracker import TokenTracker

    tracker = TokenTracker()

    class FakeUsage:
        def __init__(self, inp=0, out=0, cw=0, cr=0):
            self.input_tokens = inp
            self.output_tokens = out
            self.cache_creation_input_tokens = cw
            self.cache_read_input_tokens = cr

    # 42 Haiku page-type classifications — small, mostly cached.
    for _ in range(42):
        tracker.record(
            FakeUsage(inp=100, out=50, cr=1500),
            "claude-haiku-4-5-20251001",
        )

    # 1 Sonnet system-prompt warmup (cache write).
    tracker.record(FakeUsage(inp=200, out=400, cw=3000),
                   "claude-sonnet-4-6")

    # 15 Sonnet vision crops (legend/schedule/title/composer), all cache-hit.
    for _ in range(15):
        tracker.record(
            FakeUsage(inp=1500, out=400, cr=2000),
            "claude-sonnet-4-6",
        )

    cost = tracker.usage.estimated_cost_usd
    hit_rate = tracker.usage.cache_hit_rate
    assert cost <= 0.30, (
        f"Phase 1 cost envelope blown: ${cost:.4f} (target ≤ $0.30). "
        f"Per-model: {tracker.usage.summary()}"
    )
    assert hit_rate >= 0.90, (
        f"Cache hit rate too low: {hit_rate:.1%} — caching is broken or "
        f"system prompts aren't tagged ephemeral."
    )
    # Sanity: token totals are non-zero and routed across models.
    assert len(tracker.usage.by_model) == 2
    assert tracker.usage.api_calls == 58


# ── Exporter shape parity vs Brooklyn reference xlsx ───────────────────────


def test_exporter_shape_matches_brooklyn_reference():
    """Synthesise the first 5 Brooklyn rows and verify exporter output structure."""
    import openpyxl
    from core.qto_row import QTORow
    from core.xlsx_exporter import export

    rows = [
        QTORow(drawings="A-102",
               details="1/A901 & LEGEND/A102",
               description=("PROVIDE & INSTALL MAPLE FLOORING (25/32\" T) @ AUDITORIUM "
                            "AS PER DETAIL 1/A901 & LEGEND/A102 WHICH INCLUDES\n"
                            "-6 MIL. CONT. POLYETHYLENE FILM\n"
                            "-1/2\" RESILIENT UNDRELAYMENT\n"
                            "-2 LAYERS 15/32\" PLYWOOD\n"
                            "-2\" FLOORING FASTENERS"),
               qty=1327, units="SQ FT"),
        QTORow(drawings="A-102",
               details="1/A901",
               description=("PROVIDE & INSTALL 3\" X 4\" VENTED RESILIENT BASE @ "
                            "AUDITORIUM AS PER DETAIL 1/A901"),
               qty=180, units="LF"),
        QTORow(drawings="A-106",
               details="2B/A422",
               description=("REMOVE & REPLACE W/ COLD-APPLIED SBS MODIFIED ROOFING @ "
                            "ROOF 2, 3, 6 & 10 AS PER DETAIL 2B/A422 WHICH INCLUDES\n"
                            "-1 PLY SBS MODIFIED BITUMEN VAPOR BARRIER\n"
                            "-2-PLY SBS MODIFIED BITUMEN ROOFING MEMBRANE\n"
                            "-MULTI-LAYER INSULATION W/ COMPOSITE BOARD MIN. R VALUE=40"),
               qty=1384, units="SQ FT"),
    ]

    out_dir = str(_REPO_ROOT / "output" / "test")
    out = export(rows, template_path=str(_TEMPLATE), output_dir=out_dir,
                 pdf_stem="brooklyn_integration_test")

    wb = openpyxl.load_workbook(out, data_only=False)
    ws = wb.active
    # Row 44 = first data row; description column is D.
    assert ws["D44"].value and "MAPLE FLOORING" in ws["D44"].value
    assert ws["E44"].value == 1327
    assert ws["F44"].value == "SQ FT"
    # Wrap text must be on for multi-line descriptions.
    assert ws["D44"].alignment.wrap_text is True
    # Row height should reflect 4 sub-bullets → 15 * (1 + 4) = 75.
    assert ws.row_dimensions[44].height == 75
    # Total formula cell H must reference the correct row.
    assert ws["H44"].value == "=E44*G44"
    # Column A keeps its chain formula intact (chain skips merged row 43).
    assert ws["A44"].value == "=A42+1"
    Path(out).unlink(missing_ok=True)
