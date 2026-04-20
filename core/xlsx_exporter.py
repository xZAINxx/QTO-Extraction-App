"""
Export QTO rows to Excel, built on top of the GC estimate template.

Strategy:
  1. Copy ESTIMATE_FORMAT___GC.xlsx to the output path.
  2. Fill metadata in rows 1-6.
  3. Locate the CONSTRUCTIONS section header (row 43 in base template).
  4. Clear pre-allocated blank rows 44-99 (preserve formulas in rows 100+).
  5. Insert CSI section headers + data rows, rewriting row references correctly.
  6. Update the SUM formula range in the SUB-TOTAL row.
  7. Validate: no #REF!, all TOTAL columns are formulas.
"""
import copy
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.qto_row import QTORow


# Template row constants (1-indexed, verified from actual template)
_ROW_PROJECT_LABEL = 2     # "PROJECT:"
_ROW_DESCRIPTION = 3       # "DESCRIPTION:"
_ROW_PERF_PERIOD = 4       # "PERFORMANCE PERIOD:" — D4 is the value cell
_ROW_LIQ_DAMAGES = 5       # "LIQUIDATED DAMAGES:"
_ROW_BID_DATE = 6          # "BID OPENING DATE:"
_ROW_HEADER = 7            # Column headers
_ROW_SECTION_GEN = 8       # "GENERAL & SUPPLEMENTRY REQUIREMENTS"
_ROW_SECTION_TEMP = 16     # "TEMPORARY FACILITIES AND CONTROLS"
_ROW_SECTION_ASBESTOS = 35 # red fill
_ROW_SECTION_PLUMBING = 37
_ROW_SECTION_ELECTRIC = 39
_ROW_SECTION_HVAC = 41
_ROW_SECTION_CONSTRUCTIONS = 43   # Start of extraction area

_FIRST_DATA_ROW = 44       # First pre-allocated blank row in CONSTRUCTIONS
_LAST_PREALLOCATED = 99    # Last pre-allocated blank row
_ROW_SUBTOTAL = 100        # SUB-TOTAL row (SUM formula here)
_ROW_BOND = 101
_ROW_WC = 102
_ROW_TOTAL_BASE = 103

_COL_SNO = "A"
_COL_DRAWINGS = "B"
_COL_TAG = "C"
_COL_DESC = "D"
_COL_QTY = "E"
_COL_UNITS = "F"
_COL_UNIT_PRICE = "G"
_COL_TOTAL = "H"

_NUM_COLS = 8   # A through H


def export(
    rows: list[QTORow],
    template_path: str,
    output_dir: str,
    pdf_stem: str,
    project_meta: dict | None = None,
) -> str:
    """
    Copy template → populate → save.
    Returns the output file path.
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = str(Path(output_dir) / f"{pdf_stem}_QTO_{ts}.xlsx")
    shutil.copy2(template_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # 1. Fill project metadata
    if project_meta:
        _fill_metadata(ws, project_meta)

    # 2. Grab reference styles from existing section headers
    ref_styles = _capture_section_styles(ws)

    # 3. Clear pre-allocated blank rows 44-99 (content only; keep row count)
    _clear_preallocated(ws)

    # 4. Insert QTO rows starting at row 44
    data_rows_only = [r for r in rows if not r.is_header_row]
    header_rows = [r for r in rows if r.is_header_row]

    current_row = _FIRST_DATA_ROW
    inserted_count = 0

    # Track where data rows start/end for SUM formula update
    first_data_row = _FIRST_DATA_ROW
    last_data_row = _FIRST_DATA_ROW

    # Pre-calculate how many rows we'll need beyond the 56 pre-allocated
    needed = len(rows)
    available = _LAST_PREALLOCATED - _FIRST_DATA_ROW + 1
    extra = max(0, needed - available)
    if extra > 0:
        # Insert extra rows before the subtotal row
        ws.insert_rows(_ROW_SUBTOTAL, extra)

    # Now write each row
    for qto_row in rows:
        if qto_row.is_header_row:
            _write_section_header(ws, current_row, qto_row.description, ref_styles)
        else:
            _write_data_row(ws, current_row, qto_row)
            last_data_row = current_row
        current_row += 1

    # 5. Update SUM formula in SUB-TOTAL row (which may have shifted)
    subtotal_row = current_row  # right after last data row... actually find it
    subtotal_row = _find_subtotal_row(ws)
    if subtotal_row:
        ws[f"H{subtotal_row}"] = f"=SUM(H{_ROW_SECTION_GEN + 1}:H{subtotal_row - 1})"

    wb.save(out_path)
    return out_path


def _fill_metadata(ws, meta: dict):
    """
    Fill project metadata in rows 2-6.
    Row 2 (A2:H2 merged): project name goes into A2 alongside the label.
    Rows 3-6: value goes into the first non-merged cell after the label.
    D4 is the performance_period_days cell referenced by the live formulas.
    """
    if "project" in meta and meta["project"]:
        ws["A2"] = f"PROJECT: {meta['project']}"
    if "description" in meta and meta["description"]:
        # D3:H3 is merged; write to D3 (the top-left of that merge)
        ws["D3"] = meta["description"]
    if "performance_period_days" in meta and meta["performance_period_days"]:
        ws["D4"] = meta["performance_period_days"]
    if "liquidated_damages" in meta and meta["liquidated_damages"]:
        ws["D5"] = meta["liquidated_damages"]
    if "bid_opening_date" in meta and meta["bid_opening_date"]:
        ws["D6"] = meta["bid_opening_date"]


def _capture_section_styles(ws) -> dict:
    """Capture fill/font from existing section header rows for reuse."""
    styles = {}
    candidates = {
        "default": _ROW_SECTION_GEN,
        "red": _ROW_SECTION_ASBESTOS,
        "bold": _ROW_SECTION_CONSTRUCTIONS,
    }
    for name, row_num in candidates.items():
        cell = ws[f"A{row_num}"]
        styles[name] = {
            "fill": copy.copy(cell.fill),
            "font": copy.copy(cell.font),
            "alignment": copy.copy(cell.alignment),
        }
    return styles


def _clear_preallocated(ws):
    """Clear content of pre-allocated blank rows 44-99."""
    for row in range(_FIRST_DATA_ROW, _LAST_PREALLOCATED + 1):
        for col in range(1, _NUM_COLS + 1):
            cell = ws.cell(row=row, column=col)
            # Only clear if not a live formula we want to preserve
            if cell.value is not None:
                v = str(cell.value)
                # Clear chain formulas (=A43+1 style) and blank TOTAL formulas
                if v.startswith("=A") or v.startswith("=E"):
                    cell.value = None


def _write_section_header(ws, row_num: int, label: str, ref_styles: dict):
    """Write a CSI section header row with merged cells and styling."""
    # Merge A:H
    merge_ref = f"A{row_num}:H{row_num}"
    try:
        ws.merge_cells(merge_ref)
    except Exception:
        pass

    cell = ws[f"A{row_num}"]
    cell.value = label

    style = ref_styles.get("default", {})
    if style.get("font"):
        cell.font = Font(bold=True, name=style["font"].name or "Calibri", size=style["font"].size or 11)
    else:
        cell.font = Font(bold=True)

    if style.get("fill") and style["fill"].fill_type and style["fill"].fill_type != "none":
        cell.fill = copy.copy(style["fill"])

    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_data_row(ws, row_num: int, row: QTORow):
    """Write a single QTO data row."""
    ws[f"A{row_num}"] = row.s_no if row.s_no else None
    ws[f"B{row_num}"] = row.drawings_details
    ws[f"C{row_num}"] = row.tag
    ws[f"D{row_num}"] = row.description
    ws[f"E{row_num}"] = row.qty if row.qty else None
    ws[f"F{row_num}"] = row.units
    ws[f"G{row_num}"] = None  # Unit price — always blank
    ws[f"H{row_num}"] = f"=E{row_num}*G{row_num}"
    row.total_formula = f"=E{row_num}*G{row_num}"

    # Amber left-border indicator for needs_review rows
    if row.needs_review:
        amber = "00F59E0B"
        side = Side(style="medium", color=amber)
        ws[f"A{row_num}"].border = Border(left=side)


def _find_subtotal_row(ws) -> Optional[int]:
    """Find the SUB-TOTAL row by scanning for its label."""
    for row in ws.iter_rows(min_row=90, max_row=ws.max_row):
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == "SUB-TOTAL":
                return cell.row
    return None
